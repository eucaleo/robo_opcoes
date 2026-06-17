from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


HEADERS = [
    "codigo_opcao",
    "ativo_base",
    "call_put",
    "strike",
    "vencimento",
    "ultimo_preco",
    "ultima_quantidade",
    "bid",
    "ask",
    "volume",
    "iv",
    "delta",
    "gamma",
    "theta",
    "vega",
]

RTD = 'RTD("btg_pro_rtd","","{topic}",$A{row})'


def formula(topic: str, row: int) -> str:
    return "=" + RTD.format(topic=topic, row=row)


def parse_tickers(raw: str) -> list[str]:
    return [x.strip().upper() for x in raw.replace(";", ",").split(",") if x.strip()]


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Cria/atualiza aba RTD_OPTION_QUOTES tabular em LISTA_RTD.xlsm."
    )
    parser.add_argument("--workbook", default="LISTA_RTD.xlsm")
    parser.add_argument(
        "--tickers",
        required=True,
        help="Lista separada por vírgula. Ex: BPAC11 ou PETRA300,PETRM300",
    )
    parser.add_argument("--sheet", default="RTD_OPTION_QUOTES")
    args = parser.parse_args()

    path = Path(args.workbook)
    if not path.exists():
        raise SystemExit(f"Workbook não encontrado: {path}")

    tickers = parse_tickers(args.tickers)
    if not tickers:
        raise SystemExit("Nenhum ticker informado.")

    wb = load_workbook(path, keep_vba=True)

    if args.sheet in wb.sheetnames:
        ws = wb[args.sheet]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(args.sheet, 0)

    for col, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col).value = header

    for idx, ticker in enumerate(tickers, start=2):
        ws.cell(row=idx, column=1).value = ticker

        ws.cell(row=idx, column=2).value = formula("QUOTE.UNDERLYING_SYMBOL", idx)
        ws.cell(row=idx, column=3).value = formula("QUOTE.OPTION_TYPE", idx)
        ws.cell(row=idx, column=4).value = formula("QUOTE.STRIKE_PRICE", idx)
        ws.cell(row=idx, column=5).value = formula("QUOTE.MATURITYDATE", idx)
        ws.cell(row=idx, column=6).value = formula("QUOTE.LAST_TRADE_PRICE", idx)
        ws.cell(row=idx, column=7).value = formula("QUOTE.LAST_TRADE_QUANTITY", idx)
        ws.cell(row=idx, column=8).value = formula("QUOTE.BID_PRICE", idx)
        ws.cell(row=idx, column=9).value = formula("QUOTE.ASK_PRICE", idx)
        ws.cell(row=idx, column=10).value = formula("QUOTE.VOLUME", idx)
        ws.cell(row=idx, column=11).value = formula("QUOTE.IMPLIED_VOLATILITY", idx)
        ws.cell(row=idx, column=12).value = formula("QUOTE.DELTA", idx)
        ws.cell(row=idx, column=13).value = formula("QUOTE.GAMMA", idx)
        ws.cell(row=idx, column=14).value = formula("QUOTE.THETA", idx)
        ws.cell(row=idx, column=15).value = formula("QUOTE.VEGA", idx)

    ws.freeze_panes = "A2"

    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20

    wb.save(path)
    wb.close()

    print("[OK] Aba tabular criada/atualizada.")
    print("Workbook:", path.resolve())
    print("Sheet:", args.sheet)
    print("Tickers:", ", ".join(tickers))
    print("Linhas de dados:", len(tickers))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
